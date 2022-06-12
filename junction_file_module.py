import openpyxl
from sub_junction import Junction
from trafic_light import TrafficLight

min_row = 2
junction_name_col = 2
number_of_traffic_lights_col = 3
junction_path_col = 4


def set_data(sheet):
    row = min_row
    junction_name = sheet.cell(min_row, junction_name_col).value
    num_of_traffic_light_in_junction = sheet.cell(min_row, number_of_traffic_lights_col).value
    row = min_row+1
    num_of_traffic_lights_at_path = sheet.cell(row, junction_path_col).value
    junction_path = sheet.cell(row, junction_path_col).value
    traffic_light_name = sheet.cell(row, junction_path_col+1).value
    traffic_light_name_data = sheet.cell(row, junction_path_col+2).value
    #print(traffic_light_name_data)
    turn = traffic_light_name_data.split(',')[0].split("=")[0]
    cd = traffic_light_name_data.split(',')[0].split("=")[1]

    a = {
        turn: cd
    }
    #print(a)


def open_file():
    junction_file = openpyxl.load_workbook("Junction_Info.xlsx")
    sheet1 = junction_file["Sheet1"]
    return sheet1


def extract_junction_info():
    sheet = open_file()
    set_data(sheet)


try:
    extract_junction_info()

except FileExistsError as error:
    print(error)
except EOFError as error:
    print(error)
except FileNotFoundError as error:
    print(error)



